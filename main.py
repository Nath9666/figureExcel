import pandas as pd
import sys
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import ScatterChart, Reference, Series
import configparser
import os


def lire_config(chemin_config):
    config = configparser.ConfigParser()
    config.read(chemin_config, encoding='utf-8')
    params = config['Graphique']
    largeur = int(params.get('largeur', 800))
    hauteur = int(params.get('hauteur', 600))
    arrondi = int(params.get('arrondi', 1))
    couleur_ligne = params.get('couleur_ligne', 'blue')
    return largeur, hauteur, arrondi, couleur_ligne


def lire_donnees_excel(chemin_fichier):
    try:
        df = pd.read_excel(chemin_fichier)
        print(df)
        return df
    except Exception as e:
        print(f"Erreur lors de la lecture du fichier : {e}")
        return None


def generer_courbe(df, largeur, hauteur, arrondi, couleur_ligne, out_img):
    # On suppose que les colonnes sont 'Date' et 'Données 1'
    df['Date'] = pd.to_datetime(df['Date'])
    df = df.sort_values('Date')
    df['Données 1'] = df['Données 1'].round(arrondi)
    plt.figure(figsize=(largeur/100, hauteur/100))
    plt.plot(df['Date'], df['Données 1'], color=couleur_ligne, marker='o')
    plt.xlabel('Date')
    plt.ylabel('Données 1')
    plt.title('Courbe XY Données 1')
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(out_img)
    plt.close()


def inserer_courbe_excel(df, out_xlsx):
    # Crée un fichier Excel avec les données et ajoute une courbe XY (Scatter)
    wb = Workbook()
    ws = wb.active
    ws.title = "Données"
    # Ecrit les en-têtes
    ws.append(list(df.columns))
    # Ecrit les données
    for row in df.itertuples(index=False):
        ws.append(list(row))
    # Ajoute la courbe XY

    chart = ScatterChart()
    chart.title = "Courbe XY Données 1"
    chart.x_axis.title = 'Date'
    chart.y_axis.title = 'Données 1'
    # Les dates sont en colonne 1, les données en colonne 2
    xvalues = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    yvalues = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    series = Series(yvalues, xvalues, title_from_data=False)
    chart.series.append(series)
    ws.add_chart(chart, "E2")
    wb.save(out_xlsx)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage : python main.py <chemin_fichier_excel>")
    else:
        chemin_excel = sys.argv[1]
        chemin_config = "config.ini"
        out_xlsx = "out.xlsx"
        largeur, hauteur, arrondi, couleur_ligne = lire_config(chemin_config)
        df = lire_donnees_excel(chemin_excel)
        if df is not None:
            df['Date'] = pd.to_datetime(df['Date'])
            df = df.sort_values('Date')
            df['Données 1'] = df['Données 1'].round(arrondi)
            inserer_courbe_excel(df, out_xlsx)
            print(f"Courbe Excel générée dans {out_xlsx}")
        else:
            print("Erreur lors de la génération de la courbe.")